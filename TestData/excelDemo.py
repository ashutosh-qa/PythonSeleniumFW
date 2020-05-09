import openpyxl
# Read data from excel
book = openpyxl.load_workbook("C:\\Users\\ashut\\Documents\\DataDriven.xlsx")
sheet = book.active  # active for active sheet
Dict = {}  # Dict: <class 'dict'>: {'Fname': 'Ashutosh', 'Lname': 'Singh'}

cell = sheet.cell(row=1, column=2)
print(cell.value)

#Write data to excel
sheet.cell(row=2, column=2).value="Ashutosh"
print(sheet.cell(row=2, column=2).value)

#print(sheet.max_column)
#print(sheet.max_row)
#print(sheet['A5'].value)

# How to print table
for i in range(1, sheet.max_row+1):  #To get rows
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column+1):  # To get columns
            print(sheet.cell(row=i, column=j).value)


# How to print table
for i in range(1, sheet.max_row+1):  #To get rows
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column+1):  # To get columns
            #Dict["lastname"]="Singh"
            Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i, column=j).value

print(Dict)