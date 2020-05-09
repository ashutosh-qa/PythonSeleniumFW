import openpyxl

class HomePageData:
    test_HomePage_data =[{"Fname": "Ashutosh", "Email": "ashutosh.qa@gmail.com", "Gender": "Male"},{"Fname": "Ankita", "Email": "an.qa@gmail.com", "Gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}  # Dict: <class 'dict'>: {'Fname': 'Ashutosh', 'Lname': 'Singh'}
        book = openpyxl.load_workbook("C:\\Users\\ashut\\Documents\\DataDriven.xlsx")
        sheet = book.active  # active for active sheet
        for i in range(1, sheet.max_row+1):  #To get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column+1):  # To get columns
                    #Dict["lastname"]="Singh"
                    Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i, column=j).value
        return[Dict]